import openpyxl
from datetime import datetime

WEEKDAYS_NO = ["Mandag", "Tirsdag", "Onsdag", "Torsdag", "Fredag", "Lørdag", "Søndag"]

def _normalize(s: str) -> str:
    return " ".join(str(s).strip().split())

def _find_day_start_rows(sheet):
    """
    Finds the start row for each weekday by scanning column A for weekday names.
    Returns dict: {0: row_mandag, 1: row_tirsdag, ... 4: row_fredag}
    """
    day_to_row = {}
    targets = {name.lower(): idx for idx, name in enumerate(WEEKDAYS_NO[:5])}  # Mon-Fri

    for r in range(1, sheet.max_row + 1):
        v = sheet.cell(row=r, column=1).value  # Col A
        if v is None:
            continue
        key = _normalize(v).lower().rstrip(":")
        if key in targets and targets[key] not in day_to_row:
            day_to_row[targets[key]] = r

        if len(day_to_row) == 5:
            break

    return day_to_row

def _read_day_menu(sheet, start_row: int):
    """
    Expects:
      - Column A: menu items, and weekday label on start_row.
      - Column B: 'dagens suppe' on some row in the block.
    Reads a block until the next weekday label or empty stretch.
    Returns (items: list[str], soup: str|None)
    """
    items = []
    soup = None

    r = start_row + 1  # start after weekday label

    while r <= sheet.max_row:
        a = sheet.cell(row=r, column=1).value  # Col A (dish)
        b = sheet.cell(row=r, column=2).value  # Col B (soup)

        # Stop if we hit next weekday label in col A
        if a is not None:
            a_norm = _normalize(a).lower().rstrip(":")
            if a_norm in {d.lower() for d in WEEKDAYS_NO[:5]}:
                break

        # Capture soup if present in column B
        if b is not None and str(b).strip():
            soup = _normalize(b)

        # Capture dish if present in column A
        if a is not None and str(a).strip():
            items.append(_normalize(a))

        # Heuristic stop: if both columns are empty for 3 rows, break
        if (a is None or not str(a).strip()) and (b is None or not str(b).strip()):
            empty_run = 0
            rr = r
            while rr <= sheet.max_row:
                aa = sheet.cell(row=rr, column=1).value
                bb = sheet.cell(row=rr, column=2).value
                if (aa is None or not str(aa).strip()) and (bb is None or not str(bb).strip()):
                    empty_run += 1
                    rr += 1
                    if empty_run >= 3:
                        return items, soup
                else:
                    break

        r += 1

    return items, soup

def _read_diverse_info(sheet, start_row: int = 5, col: int = 4):
    """
    Reads 'diverse info' from column D starting at row 5 and downwards.
    Stops after 3 consecutive empty cells.
    Returns list[str].
    """
    info = []
    empty_run = 0
    r = start_row

    while r <= sheet.max_row:
        v = sheet.cell(row=r, column=col).value  # D = 4
        if v is None or not str(v).strip():
            empty_run += 1
            if empty_run >= 3:
                break
        else:
            empty_run = 0
            info.append(_normalize(v))
        r += 1

    return info

def read_menu(filename: str, day: int = -1):
    """
    day:
      -1 = today
       0-4 = Monday-Friday
    """
    wb = openpyxl.load_workbook(filename, data_only=True)
    if "Meny" not in wb.sheetnames:
        raise ValueError("Fant ikke arket 'Meny' i Excel-fila.")

    sheet = wb["Meny"]

    # Resolve day
    if day == -1:
        day = datetime.now().weekday()
        if day > 4:
            print("Ingen meny på lørdager og søndager. Kom tilbake på mandag, eller velg ukedag.")
            return

    # Locate weekday start rows dynamically
    day_start_rows = _find_day_start_rows(sheet)
    missing = [WEEKDAYS_NO[i] for i in range(5) if i not in day_start_rows]
    if missing:
        raise ValueError(
            f"Fant ikke start-rad for: {', '.join(missing)}. Sjekk at kolonne A inneholder ukedag-navnene."
        )

    # Read all weekdays (Mon-Fri)
    menus = {}
    for i in range(5):
        start_row = day_start_rows[i]
        items, soup = _read_day_menu(sheet, start_row)
        items = [x for x in items if _normalize(x).lower().rstrip(":") != WEEKDAYS_NO[i].lower()]
        menus[i] = (items, soup)

    # Read diverse info from D5 and down
    diverse_info = _read_diverse_info(sheet, start_row=5, col=4)

    # Header: Today's lunch
    today_name = WEEKDAYS_NO[day]
    today_date = datetime.today().strftime("%d.%m.%Y")
    items, soup = menus[day]

    print(f"Dagens lunsj - {today_name} {today_date}")
    for it in items[:2]:
        print(it)
    print(f"Dagens suppe: {soup}" if soup else "Dagens suppe: (ikke oppgitt)")
    print()

    # Insert diverse info between today's menu and rest of week
    if diverse_info:
        print("Diverse info:")
        for line in diverse_info:
            print(line)
        print()

    # Full week
    for i in range(5):
        day_name = WEEKDAYS_NO[i]
        items, soup = menus[i]
        print(f"{day_name}:")
        for it in items[:2]:
            print(it)
        print(soup if soup else "(ikke oppgitt)")
        print()

# Example usage
filename = "Lunsj_Gjovik.xlsx"
read_menu(filename, day=-1)
